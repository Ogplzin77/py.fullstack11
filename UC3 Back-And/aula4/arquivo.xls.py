from openpyxl import Workbook # Para instalar: pip install openpyxl

# Criando uma planilha Excel
wb = Workbook()
ws = wb.active
ws.title = "Funcionários" # Definindo o nome da planilha 

# Adicionando dados
ws.append(["Nome", "Cargo", "Salário"]) # Adicionando uma linha com os nomes das colunas 
ws.append(["Maria", "Engenheira", 8500])
ws.append(["Pedro", "Analista", 6000])

# Salvando o arquivo
wb.save("funcionarios.xlsx") # Salva o arquivo com o nome "funcionarios.xlsx"

# Lendo o arquivo
from openpyxl import load_workbook  #load_workbook é uma função do módulo openpyxl para carregar um arquivo Excel em um objeto Python workbook usando o módulo openpyxl. 
wb = load_workbook("funcionarios.xlsx") # Carregando o arquivo 
ws = wb.active # Selecionando a planilha ativa do arquivo 

# Iterando as linhas do arquivo
for linha in ws.iter_rows(values_only=True): # Iterando as linhas do arquivo e exibindo os valores das celulas da linha atual 
    print(linha)

#EXTRA: lendo xlsx com pandas
import pandas as pd # Para instalar: pip install pandas

df = pd.read_excel('funcionarios.xlsx') # Carregando o arquivo 
print(df.head())  # Exibe as primeiras linhas do arquivo 